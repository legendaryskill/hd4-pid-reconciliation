"""
Build the reconciliation output spreadsheet with formatting and conditional highlighting.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Shared styles
_HDR_FONT = Font(name='Arial', bold=True, size=10, color='FFFFFF')
_DATA_FONT = Font(name='Arial', size=9)
_BORDER = Border(
    left=Side('thin', 'B4C6E7'), right=Side('thin', 'B4C6E7'),
    top=Side('thin', 'B4C6E7'), bottom=Side('thin', 'B4C6E7'),
)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Status-based row fills
_STATUS_FILLS = {
    'OK': PatternFill('solid', fgColor='E2EFDA'),
    'MISMATCH': PatternFill('solid', fgColor='FCE4EC'),
    'CHECK': PatternFill('solid', fgColor='FFF9C4'),
    'MISSING': PatternFill('solid', fgColor='FCE4EC'),
    'TAG MISMATCH': PatternFill('solid', fgColor='FFF3E0'),
    'INSTRUMENT': PatternFill('solid', fgColor='E3F2FD'),
    'EXTRA': PatternFill('solid', fgColor='FFF3E0'),
}


def _get_status_fill(status_value):
    s = str(status_value).upper()
    for key, fill in _STATUS_FILLS.items():
        if key in s:
            return fill
    return None


def _write_df(ws, df, hdr_color, title, subtitle=''):
    """Write a DataFrame to a worksheet with formatting."""
    ws.merge_cells('A1:K1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Arial', bold=True, size=13, color=hdr_color)
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells('A2:K2')
        ws['A2'] = subtitle
        ws['A2'].font = Font(name='Arial', size=9, italic=True, color='666666')

    hdr_fill = PatternFill('solid', fgColor=hdr_color)
    cols = list(df.columns)
    sr = 4  # start row for headers

    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=sr, column=ci, value=col.replace('_', ' ').title())
        c.font = _HDR_FONT
        c.fill = hdr_fill
        c.alignment = _CENTER
        c.border = _BORDER
    ws.row_dimensions[sr].height = 32

    for ri, (_, row) in enumerate(df.iterrows(), sr + 1):
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if pd.isna(val):
                val = ''
            c = ws.cell(row=ri, column=ci, value=str(val) if val != '' else '')
            c.font = _DATA_FONT
            c.border = _BORDER
            c.alignment = _LEFT if ci <= 2 else _CENTER
            if col == 'status':
                fill = _get_status_fill(val)
                if fill:
                    c.fill = fill

    # Auto column widths
    for ci, col in enumerate(cols, 1):
        series = df[col].astype(str)
        max_len = max(len(col), int(series.str.len().max()) if len(series) > 0 else 5)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 55)

    ws.freeze_panes = f'A{sr + 1}'
    last_col = get_column_letter(len(cols))
    ws.auto_filter.ref = f'A{sr}:{last_col}{sr + len(df)}'


def build_workbook(valve_recon, equip_recon, counts, output_path):
    """
    Build the final reconciliation workbook.

    Args:
        valve_recon: DataFrame from reconcile_valves()
        equip_recon: DataFrame from reconcile_equipment()
        counts: dict with document extraction counts for the summary tab
        output_path: where to save the .xlsx
    """
    wb = Workbook()

    # --- Valve Reconciliation ---
    ws1 = wb.active
    ws1.title = "Valve Reconciliation"
    ws1.sheet_properties.tabColor = '2F5496'
    _write_df(ws1, valve_recon, '2F5496',
              'VALVE RECONCILIATION — P&ID vs VALVE LIST',
              'PN12442 HD4 Early Tonnes — Automated Cross-Reference')

    # --- Equipment Reconciliation ---
    ws2 = wb.create_sheet("Equipment Reconciliation")
    ws2.sheet_properties.tabColor = '548235'
    _write_df(ws2, equip_recon, '548235',
              'EQUIPMENT RECONCILIATION — P&ID vs EQUIPMENT LIST',
              'PN12442 HD4 Early Tonnes — Automated Cross-Reference')

    # --- Summary ---
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_properties.tabColor = 'C00000'
    ws3['A1'] = 'RECONCILIATION SUMMARY'
    ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='C00000')
    ws3.column_dimensions['A'].width = 58
    ws3.column_dimensions['B'].width = 22

    def _vc(status_contains):
        return len(valve_recon[valve_recon['status'].str.contains(status_contains, na=False)])

    def _ec(status_contains):
        return len(equip_recon[equip_recon['status'].str.contains(status_contains, na=False)])

    stats = [
        ('VALVE RECONCILIATION', '', True),
        ('Total entries', len(valve_recon), False),
        ('Matched OK', _vc('^OK$'), False),
        ('Genuine mismatches (status/size differs)', _vc('MISMATCH'), False),
        ('P&ID status not extracted (manual check needed)', _vc('CHECK'), False),
        ('Missing from valve list', _vc('MISSING FROM'), False),
        ('On valve list but not in P&ID extract', _vc('EXTRA'), False),
        ('', '', False),
        ('EQUIPMENT RECONCILIATION', '', True),
        ('Total entries', len(equip_recon), False),
        ('Matched OK', _ec('^OK$'), False),
        ('Missing from mech equipment list', _ec('^MISSING$'), False),
        ('Tag naming mismatch (PRV-00x vs BD2947.PRVxxx)', _ec('TAG MISMATCH'), False),
        ('Instrument/control items (check I&C list)', _ec('INSTRUMENT'), False),
        ('On equipment list but not in P&ID extract', _ec('EXTRA'), False),
        ('', '', False),
        ('SOURCE DOCUMENT COUNTS', '', True),
        ('Valve List — Fuel (REG-00021)', counts.get('vlv_fuel', '?'), False),
        ('Valve List — NPI/Dome (REG-00009)', counts.get('vlv_npi', '?'), False),
        ('Mech Equip — Fuel (REG-00019)', counts.get('eq_fuel', '?'), False),
        ('Mech Equip — NPI/Dome (REG-00018)', counts.get('eq_npi', '?'), False),
        ('Line List — Fuel (REG-00020)', counts.get('lines', '?'), False),
        ('Special Items (REG-00022)', counts.get('sp_items', '?'), False),
        ('Tie-ins (REG-00023)', counts.get('tieins', '?'), False),
        ('', '', False),
        ('KEY ACTIONS', '', True),
        ('1. Review genuine valve status mismatches', 'Filter Valve tab → MISMATCH', False),
        ('2. Verify "EXTRA" valves against P&ID drawings', 'Likely on dense VF-series sheets', False),
        ('3. Confirm PRV tag naming convention', 'PRV-00x vs BD2947.PRVxxx', False),
        ('4. Confirm DRYR01 vs DY01 tag convention', 'P&ID vs equip list naming', False),
        ('5. Check I&C instrument index for flagged items', 'Solenoid valves, flow switches, etc.', False),
        ('6. Walk down VF-00002, VF-00003, VF-00008', 'Dense B1 drawings, many items unreadable', False),
    ]
    for i, (label, value, is_header) in enumerate(stats, 3):
        c1 = ws3.cell(row=i, column=1, value=label)
        c2 = ws3.cell(row=i, column=2, value=value)
        if is_header:
            c1.font = Font(name='Arial', bold=True, size=11, color='2F5496')
        else:
            c1.font = Font(name='Arial', size=10)
            c2.font = Font(name='Arial', size=10)

    wb.save(output_path)
    print(f"\nOutput saved: {output_path}")
